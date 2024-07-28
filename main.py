import openpyxl

#planilha com nomes de alunos padrão específico que eu recebi
workbook = openpyxl.load_workbook('6 ano.xlsx')

sheet = workbook['Table 1']


#planilha gerada com separador_dados.py
workbook2 = openpyxl.load_workbook('AMERF_o_6_Ano_sep.xlsx')

sheet2 = workbook2['Sheet1']

# iterando sobre as linhas para preencher a planilha com nome de alunos
# com os dados da planilha gerada com separador de dados
for l1 in sheet.iter_rows(min_row=7):
    
    for n,l2 in enumerate(sheet2.iter_rows(min_row=2)):
        
        if l1[0].value == l2[1].value:
            
            for x in range(2,18):
                
                l1[x].value = l2[x].value

workbook.save('6 ano_teste.xlsx')
    


